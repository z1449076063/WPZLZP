"""
连接数据库，然后读取Excel文件中的数据，批量插入到数据库的表中
"""


from openpyxl import load_workbook
import pymysql


def sql(s):
    #打开数据库
    con = pymysql.connect(host="122.9.71.74",
                          port=61307,
                          user="root",
                          password="jwsz++321.",
                          db="jw_wpz_smart_pipe")
    #创建游标
    cur = con.cursor()
    #传入sql
    cur.execute(s)
    con.commit()
    #关闭数据库
    con.close()

"""

"""
def read_excel():
    # 打来文件
    wb = load_workbook("./file/sql.xlsx")
    # 获取要操作的sheet
    ws = wb["Sheet1"]
    # 获取最大行
    rows = ws.max_row
    # 获取最大列
    colus = ws.max_column

    list = []
    for i in range(2, rows + 1):
        list1 = []
        for j in range(1, colus + 1):
            list1.append(ws.cell(i, j).value)
        list.append(list1)
    return list


for t in read_excel():
    login_name = t[0]
    user_name = t[1]
    user_contact = t[2]
    is_enable = t[3]
    is_alone = t[4]
    role_type = t[5]
    user_pwd = t[6]
    sql(f'INSERT INTO sys_user(login_name,user_name,user_contact,is_enable,is_alone,role_type,user_pwd) VALUES ("{login_name}","{user_name}","{user_contact}","{is_enable}","{is_alone}","{role_type}","{user_pwd}")')
    print("插入sql语句完成")