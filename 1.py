# from common.sql import get_sql
#
# a = get_sql(host="122.9.71.74",
#             port=61307,
#             username="root",
#             password="jwsz++321.",
#             db="jw_wpz_smart_pipe",
#             s='SELECT * FROM `water_well` where well_name LIKE "%YSJ00%" and is_delete = 0')
#
#
# for i in range(1, 9):
#     if i == a:
#         print("正确")
#     else:
#         print("不符合")
import openpyxl

# import pdfkit
#
# # configuring pdfkit to point to our installation of wkhtmltopdf
# config = pdfkit.configuration(wkhtmltopdf=r"E:\wkhtmltopdf\bin\wkhtmltopdf.exe")
#
# # converting html file to pdf file
# pdfkit.from_file(r'C:\Users\Administrator\PycharmProjects\WPZLZP\report\20230320 145828 report.html', 'output.pdf', configuration=config)


# 指定要操作的excel文件
wb = openpyxl.load_workbook("login.xlsx")
# 获取要操作的sheet
sheet = wb["login_case"]
# 获取工作表中的最大行数
rows = sheet.max_row
# 获取工作表中的最大列数
clos = sheet.max_column

r = sheet.rows
c = sheet.columns


for i in list(r):
    case = []
    for j in i :
        case.append(j.value)

    print(case)
